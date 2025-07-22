import pandas as pd
import os
import re
import unicodedata
from difflib import get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta del archivo Excel
archivo_excel = "CDEM_con_ficheros.xlsx"
archivo_salida = archivo_excel.replace(".xlsx", "_fichero_por_autoria_coloreado.xlsx")

# Funciones de limpieza
def normalizar_apellidos(autoria):
    if not isinstance(autoria, str) or ',' not in autoria:
        return ""
    apellidos = autoria.split(',')[0]  # Solo apellidos
    apellidos = unicodedata.normalize('NFKD', apellidos).encode('ASCII', 'ignore').decode('utf-8')
    apellidos = re.sub(r'[^\w\s]', '', apellidos)  # Quitar puntuación
    partes = apellidos.strip().split()
    return ''.join([parte.capitalize() for parte in partes])

# Leer Excel
df = pd.read_excel(archivo_excel)

# Verificar columnas necesarias
for col in ['Autoría', 'Fichero', 'Año', 'Faltantes']:
    if col not in df.columns:
        raise ValueError(f"Falta la columna '{col}' en el Excel.")

df['Faltantes'] = df['Faltantes'].astype(str).str.strip()
df['Fichero'] = df['Fichero'].astype(str).str.strip()
faltantes = df['Faltantes'].dropna().unique()

# Crear referencias: Apellidos normalizados y año
referencias = []
for i, fila in df.iterrows():
    apellidos = normalizar_apellidos(fila['Autoría'])
    año = str(fila['Año']).strip() if pd.notnull(fila['Año']) else ""
    if apellidos and año:
        clave = apellidos + año
        referencias.append((clave, i))

# Emparejar faltantes y registrar actualizados
coincidencias_realizadas = []
filas_actualizadas = []

for faltante in faltantes:
    coincidencia = get_close_matches(faltante, [r[0] for r in referencias], n=1, cutoff=0.75)
    if coincidencia:
        match = coincidencia[0]
        idx = next(i for r, i in referencias if r == match)
        df.at[idx, 'Fichero'] = faltante
        df.at[idx, 'Faltantes'] = ""
        coincidencias_realizadas.append((faltante, match, idx))
        filas_actualizadas.append(idx)

# Guardar sin formato primero
df.to_excel(archivo_salida, index=False)

# Colorear filas actualizadas
wb = load_workbook(archivo_salida)
ws = wb.active

# Buscar índice de la columna Fichero
col_fichero_idx = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Fichero":
        col_fichero_idx = col
        break

# Pintar celdas
fill_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
for fila_idx in filas_actualizadas:
    ws.cell(row=fila_idx + 2, column=col_fichero_idx).fill = fill_color

# Guardar archivo final
wb.save(archivo_salida)
print(f"\nArchivo coloreado guardado en: {archivo_salida}")

if coincidencias_realizadas:
    print("\nCoincidencias realizadas:")
    for faltante, match, idx in coincidencias_realizadas:
        print(f"- {faltante} asignado a fila {idx} (clave generada: {match})")
else:
    print("\nNo se encontraron coincidencias.")